import sqlite3
import time

import openpyxl


def get_DB():
    file_name = 'clients_info.xlsx'
    wb = openpyxl.load_workbook(filename=file_name)
    sheet = wb['Лист1']
    conn = sqlite3.connect("TeleBot/data/data.db")
    c = conn.cursor()
    c.execute(
        "select user_id, name, first_name, user_name, street, house, apartment,phone,dataReq, verification, userBlock from clients")
    info = c.fetchone()
    infos = []

    for i in range(2, 3000):
        for n in range(1, 10):
            sheet.cell(row=i, column=n).value = None

    i = 2
    while info is not None:
        infos.append(info)
        info = c.fetchone()
    for info in infos:
        i += 1
        for n in range(1, len(info) + 1):
            if n == 9:
                try:
                    result = time.strftime('%d.%m.%Y %H:%M', time.localtime(int(info[n - 1])))
                except:
                    result = '-'
            else:
                result = info[n - 1]

            print(result, n)
            sheet.cell(row=i, column=n).value = result
    wb.save(file_name)
    wb.close()
    doc = open(file_name, 'rb')
    return doc


def update_column_clients(user_id, column, value):
    conn = sqlite3.connect('TeleBot/data/data.db')
    c = conn.cursor()
    c.execute(f'update clients set {column}=(?) where user_id=(?)', (value, user_id))
    conn.commit()


def get_client(user_id):
    conn = sqlite3.connect('TeleBot/data/data.db')
    c = conn.cursor()
    c.execute(
        "select user_id, name, last_name, user_name, phone, street, house, apartment, first_name,"
        "verification, userBlock from clients where user_id=(?)",
        (user_id,))
    res = c.fetchone()
    return {'user_id': res[0], 'name': res[1], 'last_name': res[2], 'user_name': res[3],
            'phone': res[4], 'street': res[5],
            'house': res[6], 'apartment': res[7], 'first_name': res[8],
            'verification': res[9], 'userBlock': res[10]} if res else None


def create_request(user_id=0, name=0, first_name=0, last_name=0, user_name=0, phone=0, street=0, house=0,
                   apartment=0, typeReq=0, pcsQuests=0, car=0, timeQuests=0, status="Очікує підтвердження"):
    conn = sqlite3.connect('TeleBot/data/data.db')
    c = conn.cursor()
    c.execute(
        "insert into requests (user_id, name, first_name, last_name, user_name, dataCreate, phone, street, house,"
        " apartment, typeReq, pcsQuests, car, timeQuests, status) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (user_id, name, first_name, last_name, user_name, int(time.time()), phone, street, house,
         apartment, typeReq, pcsQuests, car, timeQuests, status))
    conn.commit()


def del_user(user_id):
    conn = sqlite3.connect('TeleBot/data/data.db')
    c = conn.cursor()
    c.execute("delete from clients where user_id=(?)", (user_id,))
    conn.commit()


def get_requests_DB():
    file_name = 'requests_info.xlsx'
    wb = openpyxl.load_workbook(filename=file_name)
    sheet = wb['Лист1']
    conn = sqlite3.connect("TeleBot/data/data.db")
    c = conn.cursor()
    c.execute("select user_id, name, first_name, user_name, dataCreate, phone, street, house,"
              " apartment, typeReq, pcsQuests, car, timeQuests from requests")
    info = c.fetchone()
    infos = []

    for i in range(2, 3000):
        for n in range(1, 10):
            sheet.cell(row=i, column=n).value = None

    i = 2
    while info is not None:
        infos.append(info)
        info = c.fetchone()
    for info in infos:
        i += 1
        for n in range(1, len(info) + 1):
            if n == 5:
                try:
                    result = time.strftime('%d.%m.%Y %H:%M', time.localtime(int(info[n - 1])))
                except:
                    result = '-'
            else:
                result = info[n - 1]

            print(result, n)
            sheet.cell(row=i, column=n).value = result

    wb.save(file_name)
    wb.close()
    doc = open(file_name, 'rb')
    return doc


def get_msgs_for_invoice():
    msgs = []
    conn = sqlite3.connect('TeleBot/data/data.db')
    c = conn.cursor()
    c.execute('select id, user_id, typeReq, pcsQuests, timeQuests, status, car from requests where msg_status_send=0 and status!=(?)',
              ('Очікує підтвердження',))
    results = c.fetchall()
    for res in results:
        msgs.append({'id': res[0], 'user_id': res[1], 'typeReq': res[2],
                     'pcsQuests': res[3], 'timeQuests': res[4], 'status': res[5], 'car': res[6]})
    return msgs


def update_status_msgs_for_invoice():
    conn = sqlite3.connect('TeleBot/data/data.db')
    c = conn.cursor()
    c.execute('update requests set msg_status_send=(?) where msg_status_send=0 and status!=(?)',
              (1, 'Очікує підтвердження',))
    conn.commit()

